import streamlit as st
import pandas as pd
import unicodedata
import io
import openpyxl
import shutil
import tempfile
import os

st.set_page_config(page_title="Consolidador de Remuneraciones", layout="wide")
st.title("Automatización de Consolidado de Remuneraciones")

col1, col2 = st.columns(2)
with col1:
    mes_input = st.selectbox("Mes", ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"])
with col2:
    anio_input = st.number_input("Año", min_value=2020, max_value=2050, value=2026)

rut_empresa = st.text_input("RUT de la Empresa", value="76.455.680-1")
st.markdown("---")

st.markdown("### 1. Plantilla de Destino")
st.info("Sube la plantilla exacta y original descargada desde Talana (ej. PLANILLA SUBIR1.xlsx).")
archivo_plantilla = st.file_uploader("Sube la Plantilla (Excel)", type=["xlsx", "xls"], key="plantilla")

st.markdown("### 2. Datos del Mes")
st.info("Sube AQUÍ todos los archivos de Enero: Libros de Remuneraciones, Costos por Trabajador, e Informes de Haberes.")
archivos_datos = st.file_uploader("Sube los Datos Múltiples (Excel)", type=["xlsx", "xls"], accept_multiple_files=True, key="datos")

def normalize_str(s):
    if not isinstance(s, str):
        return ""
    s = s.replace(' (H)', '').replace(' (D)', '').replace(' (P)', '').replace(' *', '').strip()
    return unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8').lower()

def procesar_informe_haberes(df):
    data = []
    categoria_actual = None
    for index, row in df.iterrows():
        col0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        rut = str(row.get('Rut', '')).strip() if 'Rut' in row and pd.notna(row['Rut']) else ""
        monto = row.get('Monto', 0)
        
        if col0 and col0.lower() != 'nan' and not col0.lower().startswith('total'):
            categoria_actual = col0
            
        if rut and rut.lower() != 'nan':
            if categoria_actual:
                data.append({'Rut Trabajador': rut.replace('.', ''), 'Categoria': categoria_actual, 'Monto': monto})
                
    if not data: return pd.DataFrame()
    return pd.DataFrame(data).pivot_table(index='Rut Trabajador', columns='Categoria', values='Monto', aggfunc='sum').reset_index()

if st.button("Generar Consolidado para Talana"):
    if not archivo_plantilla or not archivos_datos:
        st.warning("Faltan archivos por subir en los pasos 1 y 2.")
    else:
        with st.spinner("Inyectando datos celda por celda en la plantilla de Talana..."):
            try:
                # 1. Leer Bases y Haberes
                df_bases = []
                df_infs = []
                
                for archivo in archivos_datos:
                    if "Libro" in archivo.name or "Costo" in archivo.name:
                        df_bases.append(pd.read_excel(archivo))
                    elif "Informe Haberes" in archivo.name:
                        df_temp = pd.read_excel(archivo, skiprows=15)
                        df_infs.append(procesar_informe_haberes(df_temp))
                
                if not df_bases:
                    st.error("Error: No se subió ningún Libro de Remuneraciones ni Costo por Trabajador.")
                    st.stop()
                    
                df_base = pd.concat(df_bases, ignore_index=True)
                df_base = df_base.groupby('Rut Trabajador').first().reset_index()
                
                df_inf = pd.concat(df_infs, ignore_index=True).groupby('Rut Trabajador').sum().reset_index() if df_infs else pd.DataFrame()
                if not df_inf.empty:
                    cols_duplicadas = [c for c in df_inf.columns if c in df_base.columns and c != 'Rut Trabajador']
                    df_inf = df_inf.drop(columns=cols_duplicadas)
                    
                df_merge = pd.merge(df_base, df_inf, on='Rut Trabajador', how='left') if not df_inf.empty else df_base.copy()
                
                # 2. Guardar temporalmente la plantilla de Talana subida
                temp_dir = tempfile.mkdtemp()
                plantilla_path = os.path.join(temp_dir, "temp_plantilla.xlsx")
                with open(plantilla_path, "wb") as f:
                    f.write(archivo_plantilla.getbuffer())
                
                # 3. Abrir Plantilla con OpenPyXL para preservar formato y estructura
                wb = openpyxl.load_workbook(plantilla_path)
                ws = wb.active
                
                # Extraer las columnas exactas que Talana puso en la fila 3
                cols_in_sheet = [cell.value for cell in ws[3]]
                
                # 4. Crear DataFrame intermedio con la estructura final
                df_final = pd.DataFrame(index=df_merge.index, columns=cols_in_sheet)
                
                # Asignar Constantes
                meses_num = {'Enero':'01','Febrero':'02','Marzo':'03','Abril':'04','Mayo':'05','Junio':'06','Julio':'07','Agosto':'08','Septiembre':'09','Octubre':'10','Noviembre':'11','Diciembre':'12'}
                
                if 'RUT EMPRESA' in df_final.columns: df_final['RUT EMPRESA'] = rut_empresa
                if 'Rut Razón Social *' in df_final.columns: df_final['Rut Razón Social *'] = rut_empresa
                if 'Año_Mes (aaaamm) *' in df_final.columns: df_final['Año_Mes (aaaamm) *'] = f"{anio_input}{meses_num[mes_input]}"
                
                # Diccionario Universal
                map_universal = {
                    'Rut Trabajador': ['Rut Trabajador', 'Rut *'],
                    'Apellido Paterno': ['Apellido Paterno'],
                    'Apellido Materno': ['Apellido Materno'],
                    'Nombres': ['Nombres'],
                    'Cargo': ['Cargo'],
                    'Sueldo Base': ['Sueldo Base (H)', 'Sueldo Base *'],
                    'N° Dias Trabajados': ['N° Dias Trabajados', 'Días Trabajados del Mes *'],
                    'N° Dias Ausentes': ['N° Dias Ausentes', 'Días de Ausentismo *'],
                    'N° Dias Licencia': ['N° Dias Licencia', 'Días de licencia médica *'],
                    'Imponible': ['Total Imponible (H)', 'Remuneración Imponible *'],
                    'Base Tributable': ['Total Tributable (afecta a impuesto) *'],
                    'Líquido': ['Sueldo Liquido *', 'Alcance Liquido'],
                    'Total Haberes': ['Remuneración Total o Sueldo Bruto *'],
                    'Haberes No Imponibles': ['Remuneración No Imponible *'],
                    'Impuesto Unico': ['Impuesto Unico (D)', 'Impuestos *'],
                    'Salud': ['Cotización Institución de Salud (D)', 'Total Isapre *'],
                    'Prevision': ['Descuento Cotización AFP (D)', 'AFP *'],
                    'Seguro de Cesantía': ['Descuento Seguro Cesantia (D)', 'Seguro Cesantía Trabajador *'],
                    'APV': ['APV', 'APV (Régimen B)'],
                    'Aporte Accidentes de Trabajo Mutual': ['Aporte Accidentes de Trabajo Mutual (P)', 'Seguro Accidente de Trabajo *'],
                    'Seguro de Cesantía Empleador': ['Seguro de Cesantía Empleador (P)', 'Seguro Cesantía Empleador *'],
                    'Seguro de Invalidez y Sobrevivencia Empleador': ['Seguro de Invalidez y Sobrevivencia Empleador (P)', 'Seguro Invalidez y Supervivencia (SIS) *'],
                    'Descuentos CCAF': ['Descuento Crédito Personal CCAF', 'Créditos Personales CCAF (D)'],
                    'Anticipos': ['Descuento Anticipo', 'Anticipos (D)'],
                    'Movilizacion': ['Movilización', 'Movilización (H)'],
                    'Colacion': ['Colación', 'Colación (H)'],
                    'Cargas Familiares': ['Asignación familiar y Maternal', 'Cargas Familiares Normales (H)']
                }
                
                # Mapeo base
                for src, dst_list in map_universal.items():
                    if src in df_merge.columns:
                        for dst in dst_list:
                            if dst in df_final.columns:
                                df_final[dst] = df_merge[src]
                                
                # Mapeo dinámico por fonética
                for col in cols_in_sheet:
                    if col is None or df_final[col].notna().any(): continue
                    clean_col = normalize_str(col)
                    
                    match = None
                    for m_col in df_merge.columns:
                        if normalize_str(m_col) == clean_col:
                            match = m_col
                            break
                    if match:
                        df_final[col] = df_merge[match]

                # Rellenar Vacíos para evitar nulos
                for col in df_final.columns:
                    if str(col).startswith('Rut') or str(col).startswith('Año'):
                        df_final[col] = df_final[col].fillna("")
                    else:
                        df_final[col] = pd.to_numeric(df_final[col], errors='coerce').fillna(0)
                
                # 5. Inyectar datos en la hoja empezando en la Fila 4
                for r_idx, row in df_final.iterrows():
                    row_num = r_idx + 4 # Empieza a escribir en la fila 4
                    for c_idx, col_name in enumerate(cols_in_sheet):
                        if col_name in df_final.columns:
                            val = row[col_name]
                            ws.cell(row=row_num, column=c_idx+1, value=val)
                            
                # 6. Descargar el archivo procesado
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.success("¡Consolidado Inyectado en Plantilla! El formato de Talana está 100% intacto.")
                st.download_button(
                    label="📥 Descargar Subida Talana Definitiva", 
                    data=output, 
                    file_name=f"Subida_Talana_{mes_input}_{anio_input}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Error procesando la información: {str(e)}")
